#!/usr/bin/env python3
"""
Generate Timesheet Tab from Categorized Timeline

Creates a new weekly timesheet tab in Project-Timesheet.xlsx from categorized timeline data.
"""

import json
import sys
import argparse
from datetime import datetime, time, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter


def load_categorized_timeline(input_file):
    """Load JSON and validate structure"""
    with open(input_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Validate required fields
    required = ['week_range', 'categorized_timeline']
    for field in required:
        if field not in data:
            raise ValueError(f"Missing required field: {field}")

    return data


def validate_categories(timeline, categories_file='work-categories.json'):
    """Ensure all categories are valid"""
    with open(categories_file, 'r', encoding='utf-8') as f:
        valid_categories = [cat['name'] for cat in json.load(f)['categories']]

    invalid = []
    for entry in timeline:
        if entry.get('category') not in valid_categories:
            invalid.append(entry)

    if invalid:
        print(f"⚠️  Warning: {len(invalid)} entries have invalid categories")
        for entry in invalid:
            print(f"  - {entry['datetime']}: '{entry.get('category')}'")

    return len(invalid) == 0


def parse_time_string(time_str):
    """Convert time string to datetime.time object"""
    if not time_str:
        return None

    # Handle various formats: "10:00 AM", "10:00:00", etc.
    time_str = time_str.strip()

    # Try 12-hour format with AM/PM
    for fmt in ['%I:%M %p', '%I:%M:%S %p', '%H:%M', '%H:%M:%S']:
        try:
            dt = datetime.strptime(time_str, fmt)
            return dt.time()
        except ValueError:
            continue

    # If all formats fail, try to parse as HH:MM
    try:
        parts = time_str.split(':')
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        return time(hour, minute, 0)
    except (ValueError, IndexError):
        print(f"⚠️  Could not parse time: {time_str}")
        return None


def group_by_date(timeline):
    """Group activities by date"""
    by_date = defaultdict(list)

    for entry in timeline:
        dt = datetime.fromisoformat(entry['datetime'])
        date_key = dt.date()
        by_date[date_key].append(entry)

    # Sort each day's activities by time
    for date in by_date:
        by_date[date].sort(key=lambda e: datetime.fromisoformat(e['datetime']))

    return by_date


def consolidate_to_rows(daily_activities):
    """
    Convert activities to Excel rows.
    Each row can have up to 2 time blocks (C-D and F-G).
    """
    rows = []

    for date, activities in sorted(daily_activities.items()):
        # Each activity gets its own row for now (simple approach)
        # Could be enhanced to consolidate same-category activities
        for activity in activities:
            row = {
                'date': date,
                'category': activity['category'],
                'start_time_1': parse_time_string(activity.get('start_time', '')),
                'end_time_1': parse_time_string(activity.get('end_time', '')),
                'start_time_2': None,
                'end_time_2': None
            }
            rows.append(row)

    return rows


def create_timesheet_tab(workbook, week_start_date, activity_rows, categories):
    """Create new sheet with timesheet data"""

    # Sheet name from Saturday date
    sheet_name = week_start_date.strftime('%Y-%m-%d')

    # Create or get sheet
    if sheet_name in workbook.sheetnames:
        print(f"⚠️  Sheet '{sheet_name}' already exists - will overwrite")
        del workbook[sheet_name]

    ws = workbook.create_sheet(sheet_name)

    # Row 2: Name
    ws['A2'] = 'Name:'
    ws['B2'] = 'Ray Morris'

    # Row 4: Headers
    headers = ['Date', 'Project / Category', 'Start Time', 'End Time',
               ' ', 'StartTime', 'EndTime', 'Hours']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(4, col_idx, header)

    # Rows 5-30: Activities
    row_idx = 5
    first_activity_date = activity_rows[0]['date'] if activity_rows else week_start_date

    for activity in activity_rows:
        # Column A: Date (with formula)
        if row_idx == 5:
            ws.cell(row_idx, 1, activity['date'])
        else:
            days_diff = (activity['date'] - first_activity_date).days
            if days_diff == 0:
                ws.cell(row_idx, 1).value = f'=A5'
            else:
                ws.cell(row_idx, 1).value = f'=A5+{days_diff}'

        # Column B: Category
        ws.cell(row_idx, 2, activity['category'])

        # Column C-D: First time block
        if activity['start_time_1']:
            ws.cell(row_idx, 3, activity['start_time_1'])
            ws.cell(row_idx, 3).number_format = 'h:mm:ss'
        if activity['end_time_1']:
            ws.cell(row_idx, 4, activity['end_time_1'])
            ws.cell(row_idx, 4).number_format = 'h:mm:ss'

        # Column F-G: Second time block (if present)
        if activity['start_time_2']:
            ws.cell(row_idx, 6, activity['start_time_2'])
            ws.cell(row_idx, 6).number_format = 'h:mm:ss'
        if activity['end_time_2']:
            ws.cell(row_idx, 7, activity['end_time_2'])
            ws.cell(row_idx, 7).number_format = 'h:mm:ss'

        # Column H: Hours formula
        formula = f'=((D{row_idx}-C{row_idx})*24)+((G{row_idx}-F{row_idx})*24)'
        ws.cell(row_idx, 8).value = formula

        row_idx += 1

    # Row 32: "Weekly totals:"
    ws.cell(32, 1, 'Weekly totals:')

    # Rows 33+: Category totals
    totals_row = 33
    for category in categories:
        # Column A: Category name
        ws.cell(totals_row, 1, category)

        # Column B: SUMIF formula
        formula = f'=ROUND(SUMIF($B$5:$B$30,A{totals_row},$H$5:$H$30), 0)'
        ws.cell(totals_row, 2).value = formula

        # Column C: Percentage formula
        first_total = 33
        last_total = 33 + len(categories) - 1
        formula = f'=B{totals_row}/SUM(B${first_total}:B${last_total})'
        ws.cell(totals_row, 3).value = formula
        ws.cell(totals_row, 3).number_format = '0.0%'

        totals_row += 1

    return ws


def print_preview(activity_rows, categories):
    """Print preview of what would be generated"""
    print("=" * 80)
    print("TIMESHEET PREVIEW")
    print("=" * 80)
    print()
    print("Activities:")
    for i, row in enumerate(activity_rows, 1):
        print(f"{i:2d}. {row['date']} | {row['category']:25s} | "
              f"{row['start_time_1']} - {row['end_time_1']}")

    print()
    print("Categories:")
    for i, cat in enumerate(categories, 1):
        print(f"{i:2d}. {cat}")
    print()


def main():
    parser = argparse.ArgumentParser(description='Generate timesheet tab from categorized timeline')
    parser.add_argument('--input', '-i', default='categorized-timeline.json',
                        help='Input categorized timeline JSON file')
    parser.add_argument('--workbook', '-w', default='Project-Timesheet.xlsx',
                        help='Output Excel workbook file')
    parser.add_argument('--categories', '-c', default='work-categories.json',
                        help='Categories definition file')
    parser.add_argument('--preview', action='store_true',
                        help='Preview mode - print to console without writing Excel')

    args = parser.parse_args()

    print("=" * 80)
    print("TIMESHEET TAB GENERATOR")
    print("=" * 80)
    print()

    # Load data
    print(f"Loading timeline from {args.input}...")
    timeline_data = load_categorized_timeline(args.input)

    # Validate categories
    print(f"Validating categories...")
    validate_categories(timeline_data['categorized_timeline'], args.categories)

    # Load category list
    with open(args.categories, 'r', encoding='utf-8') as f:
        categories = [cat['name'] for cat in json.load(f)['categories']]

    # Group by date
    print(f"Grouping {len(timeline_data['categorized_timeline'])} activities by date...")
    by_date = group_by_date(timeline_data['categorized_timeline'])

    # Consolidate to rows
    activity_rows = consolidate_to_rows(by_date)
    print(f"Consolidated to {len(activity_rows)} rows")
    print()

    if args.preview:
        # Print preview
        print_preview(activity_rows, categories)
    else:
        # Load workbook
        try:
            wb = openpyxl.load_workbook(args.workbook)
        except FileNotFoundError:
            print(f"⚠️  Workbook {args.workbook} not found - creating new workbook")
            wb = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        # Create sheet
        week_start = datetime.strptime(
            timeline_data['week_range']['start'],
            '%Y-%m-%d'
        ).date()

        print(f"Creating sheet for week starting {week_start}...")
        ws = create_timesheet_tab(wb, week_start, activity_rows, categories)

        # Save
        wb.save(args.workbook)

        print()
        print(f"✓ Created sheet '{week_start.strftime('%Y-%m-%d')}' in {args.workbook}")
        print(f"  {len(activity_rows)} activities")
        print(f"  {len(categories)} categories")
        print()


if __name__ == '__main__':
    main()
